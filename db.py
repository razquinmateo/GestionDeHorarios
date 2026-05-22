import sqlite3
import os
import sys
from datetime import date
from tkinter import messagebox
from openpyxl import load_workbook

from constantes import DB_FILE, DATA_START, COLS


def db_connect():
    base = os.path.dirname(
        sys.executable if getattr(sys, "frozen", False) else os.path.abspath(__file__)
    )
    return sqlite3.connect(os.path.join(base, DB_FILE))


def db_inicializar():
    con = db_connect()
    con.execute("""
        CREATE TABLE IF NOT EXISTS empleados (
            id     INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT NOT NULL UNIQUE,
            activo INTEGER NOT NULL DEFAULT 1
        )
    """)
    con.execute("""
        CREATE TABLE IF NOT EXISTS planillas (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            empleado_id   INTEGER NOT NULL REFERENCES empleados(id),
            mes           INTEGER NOT NULL,
            anio          INTEGER NOT NULL,
            horas         TEXT DEFAULT '',
            hrs_extras    TEXT DEFAULT '',
            hrs_noct      TEXT DEFAULT '',
            adelanto      TEXT DEFAULT '',
            observaciones TEXT DEFAULT '',
            origen        TEXT DEFAULT 'manual',
            UNIQUE(empleado_id, mes, anio)
        )
    """)
    try:
        con.execute("ALTER TABLE planillas ADD COLUMN origen TEXT DEFAULT 'manual'")
    except Exception:
        pass
    try:
        con.execute("ALTER TABLE planillas ADD COLUMN horas_base TEXT DEFAULT ''")
    except Exception:
        pass
    try:
        con.execute("ALTER TABLE planillas ADD COLUMN hrs_extras_base TEXT DEFAULT ''")
    except Exception:
        pass
    try:
        con.execute("ALTER TABLE planillas ADD COLUMN hrs_noct_base TEXT DEFAULT ''")
    except Exception:
        pass
    con.execute("""
        CREATE TABLE IF NOT EXISTS registro_diario (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            empleado_id INTEGER NOT NULL REFERENCES empleados(id),
            fecha       TEXT NOT NULL,
            horas       REAL DEFAULT 0,
            hrs_extras  REAL DEFAULT 0,
            hrs_noct    REAL DEFAULT 0,
            UNIQUE(empleado_id, fecha)
        )
    """)
    try:
        con.execute("ALTER TABLE registro_diario ADD COLUMN hrs_extras REAL DEFAULT 0")
    except Exception:
        pass
    try:
        con.execute("ALTER TABLE registro_diario ADD COLUMN hrs_noct REAL DEFAULT 0")
    except Exception:
        pass
    con.commit()
    con.close()


def db_importar_excel(path):
    con = db_connect()
    cur = con.cursor()
    if cur.execute("SELECT COUNT(*) FROM empleados").fetchone()[0] > 0:
        con.close()
        return
    wb = load_workbook(path)
    ws = wb.active
    hoy = date.today()
    for row in ws.iter_rows(min_row=DATA_START, values_only=True):
        nombre = row[0]
        if nombre is None:
            continue
        nombre = str(nombre).strip()
        cur.execute("INSERT OR IGNORE INTO empleados (nombre) VALUES (?)", (nombre,))
        emp_id = cur.execute(
            "SELECT id FROM empleados WHERE nombre=?", (nombre,)
        ).fetchone()[0]
        cur.execute(
            """
            INSERT OR IGNORE INTO planillas
                (empleado_id, mes, anio, horas, hrs_extras, hrs_noct, adelanto, observaciones)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
            (
                emp_id,
                hoy.month,
                hoy.year,
                str(row[1]).strip() if row[1] is not None else "",
                str(row[2]).strip() if row[2] is not None else "",
                str(row[3]).strip() if row[3] is not None else "",
                str(row[4]).strip() if row[4] is not None else "",
                str(row[5]).strip() if row[5] is not None else "",
            ),
        )
    con.commit()
    con.close()


def db_asegurar_mes(mes, anio):
    con = db_connect()
    cur = con.cursor()
    empleados = cur.execute("SELECT id FROM empleados WHERE activo=1").fetchall()
    for (emp_id,) in empleados:
        cur.execute(
            """
            INSERT OR IGNORE INTO planillas
                (empleado_id, mes, anio, horas, hrs_extras, hrs_noct, adelanto, observaciones)
            VALUES (?, ?, ?, '', '', '', '', '')
        """,
            (emp_id, mes, anio),
        )
    con.commit()
    con.close()


def db_cargar_mes(mes, anio):
    db_asegurar_mes(mes, anio)
    con = db_connect()
    cur = con.cursor()
    rows = cur.execute(
        """
        SELECT e.nombre, p.horas, p.hrs_extras, p.hrs_noct, p.adelanto, p.observaciones, p.origen
        FROM planillas p
        JOIN empleados e ON e.id = p.empleado_id
        WHERE p.mes=? AND p.anio=? AND e.activo=1
        ORDER BY e.id
    """,
        (mes, anio),
    ).fetchall()
    con.close()
    return [
        {
            "Empleado": r[0],
            "Horas": r[1] or "",
            "Hrs. Extras": r[2] or "",
            "Hrs. Noct.": r[3] or "",
            "Adelanto": r[4] or "",
            "Observaciones": r[5] or "",
            "Origen": r[6] or "manual",
        }
        for r in rows
    ]


def db_guardar_fila(nombre, mes, anio, datos):
    con = db_connect()
    cur = con.cursor()
    emp = cur.execute("SELECT id FROM empleados WHERE nombre=?", (nombre,)).fetchone()
    if emp is None:
        con.close()
        return

    # Convertir a float las entradas manuales para almacenarlas como "base"
    def _to_float(v):
        try:
            return float(str(v).replace(" ", "").replace(",", "."))
        except Exception:
            return 0.0

    horas_base_val = _to_float(datos.get("Horas", ""))
    extras_base_val = _to_float(datos.get("Hrs. Extras", ""))
    noct_base_val = _to_float(datos.get("Hrs. Noct.", ""))

    cur.execute(
        """
        UPDATE planillas SET
            horas=?, hrs_extras=?, hrs_noct=?, adelanto=?, observaciones=?,
            origen='manual', horas_base=?, hrs_extras_base=?, hrs_noct_base=?
        WHERE empleado_id=? AND mes=? AND anio=?
    """,
        (
            datos["Horas"],
            datos["Hrs. Extras"],
            datos["Hrs. Noct."],
            datos["Adelanto"],
            datos["Observaciones"],
            # guardar como string vacío si no hay base numérica
            (
                str(int(horas_base_val))
                if horas_base_val == int(horas_base_val) and horas_base_val != 0
                else (str(horas_base_val) if horas_base_val != 0 else "")
            ),
            (
                str(int(extras_base_val))
                if extras_base_val == int(extras_base_val) and extras_base_val != 0
                else (str(extras_base_val) if extras_base_val != 0 else "")
            ),
            (
                str(int(noct_base_val))
                if noct_base_val == int(noct_base_val) and noct_base_val != 0
                else (str(noct_base_val) if noct_base_val != 0 else "")
            ),
            emp[0],
            mes,
            anio,
        ),
    )
    # Al guardar manualmente la fila mensual, eliminar registros diarios
    # de ese empleado correspondientes al mismo mes (no otros meses).
    cur.execute(
        """
        DELETE FROM registro_diario
        WHERE empleado_id=? AND strftime('%m', fecha)=? AND strftime('%Y', fecha)=?
    """,
        (emp[0], f"{mes:02d}", str(anio)),
    )
    con.commit()
    con.close()


def db_agregar_empleado(nombre, mes, anio):
    con = db_connect()
    cur = con.cursor()
    try:
        cur.execute("INSERT INTO empleados (nombre) VALUES (?)", (nombre,))
        emp_id = cur.lastrowid
        meses_existentes = cur.execute(
            "SELECT DISTINCT mes, anio FROM planillas"
        ).fetchall()
        for m, a in meses_existentes:
            cur.execute(
                """
                INSERT OR IGNORE INTO planillas
                    (empleado_id, mes, anio, horas, hrs_extras, hrs_noct, adelanto, observaciones)
                VALUES (?, ?, ?, '', '', '', '', '')
            """,
                (emp_id, m, a),
            )
        cur.execute(
            """
            INSERT OR IGNORE INTO planillas
                (empleado_id, mes, anio, horas, hrs_extras, hrs_noct, adelanto, observaciones)
            VALUES (?, ?, ?, '', '', '', '', '')
        """,
            (emp_id, mes, anio),
        )
        con.commit()
        return True
    except sqlite3.IntegrityError:
        messagebox.showwarning(
            "Duplicado", f"Ya existe un empleado llamado '{nombre}'."
        )
        return False
    finally:
        con.close()


def db_eliminar_empleado(nombre):
    con = db_connect()
    con.execute("UPDATE empleados SET activo=0 WHERE nombre=?", (nombre,))
    con.commit()
    con.close()


def db_cargar_empleados_activos():
    con = db_connect()
    rows = con.execute(
        "SELECT nombre FROM empleados WHERE activo=1 ORDER BY id"
    ).fetchall()
    con.close()
    return [r[0] for r in rows]


def _columna_registro(tipo):
    return tipo if tipo in ("horas", "hrs_extras", "hrs_noct") else "horas"


def db_cargar_registro_diario(nombre, mes, anio, columna="horas"):
    columna = _columna_registro(columna)
    con = db_connect()
    cur = con.cursor()
    emp = cur.execute("SELECT id FROM empleados WHERE nombre=?", (nombre,)).fetchone()
    if emp is None:
        con.close()
        return {}
    rows = cur.execute(
        f"""
        SELECT fecha, {columna} FROM registro_diario
        WHERE empleado_id=? AND strftime('%m', fecha)=? AND strftime('%Y', fecha)=?
    """,
        (emp[0], f"{mes:02d}", str(anio)),
    ).fetchall()
    con.close()
    return {r[0]: r[1] or 0 for r in rows}


def db_guardar_registro_dia(nombre, fecha_str, columna, horas):
    columna = _columna_registro(columna)
    con = db_connect()
    cur = con.cursor()
    emp = cur.execute("SELECT id FROM empleados WHERE nombre=?", (nombre,)).fetchone()
    if emp is None:
        con.close()
        return

    # Leer valores actuales para preservarlos
    existing = cur.execute(
        "SELECT horas, hrs_extras, hrs_noct FROM registro_diario WHERE empleado_id=? AND fecha=?",
        (emp[0], fecha_str),
    ).fetchone()

    if existing is None:
        # Nueva fila: inicializar con ceros y el valor en la columna especificada
        valores = {
            "horas": 0.0,
            "hrs_extras": 0.0,
            "hrs_noct": 0.0,
        }
        valores[columna] = horas
    else:
        # Fila existente: preservar valores actuales
        valores = {
            "horas": float(existing[0]) if existing[0] is not None else 0.0,
            "hrs_extras": float(existing[1]) if existing[1] is not None else 0.0,
            "hrs_noct": float(existing[2]) if existing[2] is not None else 0.0,
        }
        valores[columna] = horas

    # Si todos los valores son 0, eliminar el registro
    if all(v == 0.0 for v in valores.values()):
        cur.execute(
            "DELETE FROM registro_diario WHERE empleado_id=? AND fecha=?",
            (emp[0], fecha_str),
        )
    else:
        # Insertar o actualizar con todos los valores preservados
        cur.execute(
            f"""
            INSERT INTO registro_diario (empleado_id, fecha, horas, hrs_extras, hrs_noct)
            VALUES (?, ?, ?, ?, ?)
            ON CONFLICT(empleado_id, fecha) DO UPDATE SET 
                horas=excluded.horas, 
                hrs_extras=excluded.hrs_extras, 
                hrs_noct=excluded.hrs_noct
            """,
            (
                emp[0],
                fecha_str,
                valores["horas"],
                valores["hrs_extras"],
                valores["hrs_noct"],
            ),
        )
    con.commit()
    con.close()


def db_sincronizar_total_diario(nombre, mes, anio):
    """Suma horas_base + registro diario y escribe el total en horas."""
    con = db_connect()
    cur = con.cursor()
    emp = cur.execute("SELECT id FROM empleados WHERE nombre=?", (nombre,)).fetchone()
    if emp is None:
        con.close()
        return

    # Verificar si hay datos en registro_diario
    registros = cur.execute(
        """
        SELECT SUM(horas), SUM(hrs_extras), SUM(hrs_noct)
        FROM registro_diario
        WHERE empleado_id=? AND strftime('%m', fecha)=? AND strftime('%Y', fecha)=?
    """,
        (emp[0], f"{mes:02d}", str(anio)),
    ).fetchone()

    total_diario = float(registros[0] or 0)
    total_extras = float(registros[1] or 0)
    total_noct = float(registros[2] or 0)

    # Si hay datos en registro_diario, usarlos y marcar como 'diario'
    if total_diario > 0 or total_extras > 0 or total_noct > 0:

        fila = cur.execute(
            "SELECT horas_base, hrs_extras_base, hrs_noct_base FROM planillas WHERE empleado_id=? AND mes=? AND anio=?",
            (emp[0], mes, anio),
        ).fetchone()

        horas_base = 0.0
        extras_base = 0.0
        noct_base = 0.0
        if fila:
            if fila[0]:
                try:
                    horas_base = float(fila[0])
                except Exception:
                    horas_base = 0.0
            if len(fila) > 1 and fila[1]:
                try:
                    extras_base = float(fila[1])
                except Exception:
                    extras_base = 0.0
            if len(fila) > 2 and fila[2]:
                try:
                    noct_base = float(fila[2])
                except Exception:
                    noct_base = 0.0

        total_final = horas_base + total_diario
        total_extras_final = extras_base + total_extras
        total_noct_final = noct_base + total_noct

        def _format(valor):
            if valor == 0:
                return ""
            return str(int(valor)) if valor == int(valor) else str(valor)

        cur.execute(
            """
            UPDATE planillas SET horas=?, hrs_extras=?, hrs_noct=?, origen='diario', horas_base=?, hrs_extras_base=?, hrs_noct_base=?
            WHERE empleado_id=? AND mes=? AND anio=?
        """,
            (
                _format(total_final),
                _format(total_extras_final),
                _format(total_noct_final),
                str(horas_base) if horas_base > 0 else "",
                str(extras_base) if extras_base > 0 else "",
                str(noct_base) if noct_base > 0 else "",
                emp[0],
                mes,
                anio,
            ),
        )

    con.commit()
    con.close()
