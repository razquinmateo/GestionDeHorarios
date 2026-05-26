import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from constantes import COLS, MESES


def excel_exportar(path, empleados, mes, anio):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{MESES[mes-1]} {anio}"

    color_header = "2B5BA8"
    color_fila_par = "EEF2FB"
    fuente_titulo = Font(name="Segoe UI", size=14, bold=True, color="2B5BA8")
    fuente_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    fuente_normal = Font(name="Segoe UI", size=11)
    relleno_header = PatternFill("solid", fgColor=color_header)
    relleno_par = PatternFill("solid", fgColor=color_fila_par)
    alin_centro = Alignment(horizontal="center", vertical="center")
    alin_izq = Alignment(horizontal="left", vertical="center", wrap_text=True)
    borde_fino = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    ws.merge_cells("A1:F1")
    ws["A1"].value = f"Planilla de sueldos — {MESES[mes-1]} {anio}"
    ws["A1"].font = fuente_titulo
    ws["A1"].alignment = alin_izq
    ws.row_dimensions[1].height = 28

    for c, nombre_col in enumerate(COLS, start=1):
        celda = ws.cell(row=2, column=c, value=nombre_col)
        celda.font = fuente_header
        celda.fill = relleno_header
        celda.alignment = alin_centro
        celda.border = borde_fino
    ws.row_dimensions[2].height = 22

    for i, emp in enumerate(empleados):
        fila = 3 + i
        es_par = i % 2 == 0
        for c_idx, col in enumerate(COLS, start=1):
            val = emp[col]
            if col not in ("Empleado", "Observaciones") and val not in ("", "-"):
                try:
                    val = float(val) if "." in val else int(val)
                except (ValueError, TypeError):
                    pass
            celda = ws.cell(
                row=fila, column=c_idx, value=val if val not in ("", "-") else None
            )
            celda.font = fuente_normal
            celda.border = borde_fino
            celda.alignment = (
                alin_izq if col in ("Empleado", "Observaciones") else alin_centro
            )
            if es_par:
                celda.fill = relleno_par
        ws.row_dimensions[fila].height = 20

    anchos = {
        "Empleado": 24,
        "Horas": 10,
        "Hrs. Extras": 13,
        "Hrs. Noct.": 13,
        "Adelanto": 13,
        "Observaciones": 40,
    }
    for c_idx, col in enumerate(COLS, start=1):
        ws.column_dimensions[get_column_letter(c_idx)].width = anchos.get(col, 14)

    ws.freeze_panes = "A3"
    wb.save(path)


def excel_importar(path):
    wb = load_workbook(path)
    ws = wb.active
    ALIAS = {
        "empleado": "Empleado",
        "nombre": "Empleado",
        "horas": "Horas",
        "hrs. extras": "Hrs. Extras",
        "extras": "Hrs. Extras",
        "horas extra": "Hrs. Extras",
        "hrs. noct.": "Hrs. Noct.",
        "nocturnas": "Hrs. Noct.",
        "noct": "Hrs. Noct.",
        "adelanto": "Adelanto",
        "observaciones": "Observaciones",
        "obs": "Observaciones",
    }
    col_map = {}
    header_found = False
    data_rows = []

    for row in ws.iter_rows(values_only=True):
        celdas = [str(c).strip() if c is not None else "" for c in row]
        if not header_found:
            textos_lower = [c.lower() for c in celdas]
            if any(k in textos_lower for k in ("empleado", "nombre", "horas")):
                for i, celda in enumerate(textos_lower):
                    if celda in ALIAS:
                        col_map[ALIAS[celda]] = i
                header_found = True
            continue
        nombre_idx = col_map.get("Empleado", 0)
        nombre = celdas[nombre_idx] if nombre_idx < len(celdas) else ""
        if not nombre:
            continue
        emp = {"Empleado": nombre}
        for col in COLS[1:]:
            idx = col_map.get(col)
            emp[col] = (
                celdas[idx]
                if (idx is not None and idx < len(celdas) and celdas[idx])
                else "-"
            )
        data_rows.append(emp)

    return data_rows


def calcular_totales(empleados):
    total_horas = total_extra = total_adel = 0
    for emp in empleados:
        for campo in ["Horas", "Hrs. Extras", "Adelanto"]:
            val = emp[campo].replace(" ", "")
            try:
                num = float(val)
                if campo == "Horas":
                    total_horas += num
                elif campo == "Hrs. Extras":
                    total_extra += num
                else:
                    total_adel += num
            except ValueError:
                pass
    return {
        "empleados": len(empleados),
        "horas": total_horas,
        "extras": total_extra,
        "adelantos": total_adel,
    }
